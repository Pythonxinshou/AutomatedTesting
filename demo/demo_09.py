# excel文件读取
# xlrd xlwt pandas openpyxl

# openpyxl操作excel
# 1.安装  pip install openpyxl 还可以在setting
# 创建一个excel
# 1.打开excel  open with open

import openpyxl

# 加载工作簿  返回一个excel对象  工作簿在哪
wb=openpyxl.load_workbook('./data/bb.xlsx')
print(wb)
# 指定sheet表单  你的数据写在那个表单里面
sheet=wb['Sheet1']
# print(sheet)
# 拿单元格  值
# id1=sheet['A1'].value
# print(id1)
# name=sheet.cell(row=1,column=2).value
# print(name)

# 拿到所有的值  如果没有值就会是none  wb['Sheet1'].value
# values=sheet.values
# for i in values:
#     # 放在元祖中的
#     print(i)
# range左闭右开1,2,3,4,5,6
# 总共有条数据
# rows=sheet.max_row
# print(rows)
# for i in range(1,rows+1):
#     id=sheet.cell(row=i,column=1).value
#     name=sheet.cell(row=i,column=2).value
#     age=sheet.cell(row=i,column=3).value
#     sex=sheet.cell(row=i,column=4).value
#     print(id,name,age,sex)

# 读取所有的sheet
# sheets=wb.sheetnames
# # print(sheets)
# # # 读取到所有的表单  所有的表单数据
# for i in sheets:
#     print(i)
    # wb['Sheet1'].values   wb['Sheet2'].value
    # for j in wb[i].values:
    #     print(j)

# 写入数据保存在表单里面
# sheet.cell(row=1,column=5).value='结果'
# wb.save('./data/bb.xlsx')