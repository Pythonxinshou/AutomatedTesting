# -*- coding:utf-8 -*-
# @Time : 2021/11/15 11:51 
# @Author : 朱树剑
# @File : demo09.py
# @Software: PyCharm
import openpyxl

# 记载工作簿
wb = openpyxl.load_workbook(r'bb.xlsx')
print(wb)
# 指定sheet表单
sheet = wb['Sheet1']
print(sheet)
# 拿单元格
id1 = sheet['A1'].value
print(id1)
name = sheet.cell(row=1, column=2).value
print(name)

# 拿到所有值,没有值的地方返回none
values = sheet.values
for i in values:
    # 放在元祖里
    print(i)

# sheet.max_row     统计共有多少条数据
rows = sheet.max_row
print(rows)
for i in range(1, rows + 1):
    id = sheet.cell(row=i, column=1).value
    name = sheet.cell(row=i, column=2).value
    age = sheet.cell(row=i, column=3).value
    sex = sheet.cell(row=i, column=4).value
    print(id, name, age, sex)

# 读取所有的sheet
sheets = wb.sheetnames
print(sheets)
for i in sheets:
    print(i)
    for j in wb[i].values:
        print(j)

# 写入数据保存到表单里面
sheet.cell(row=1, column=5).value = '结果'
wb.save(r'./bb.xlsx')
