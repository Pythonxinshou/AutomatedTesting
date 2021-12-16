# -*- coding:utf-8 -*-
# @Time : 2021/12/13 12:00 
# @Author : 朱树剑
# @File : read.py 
# @Software: PyCharm

import openpyxl
# 读取excel中的用例正文
from Web_Keys.web_keys.keys import Keys


# r'../data/excel_driver.xlsx'
def read(file):
    excel = openpyxl.load_workbook(file)

    for name in excel.sheetnames:
        sheet = excel[name]
        print(f'*******************{name}**************')
        for values in sheet.values:
            # 第一个单元格为int类型，则表示进入测试用例正文
            if type(values[0]) is int:
                # print(values)
                # 清楚参数字典为None的参数键值对
                data = dict()
                data['name'] = values[2]
                data['value'] = values[3]
                data['text'] = values[4]
                data['expect'] = values[6]
                # print(data)
                for key in list(data.keys()):
                    if data[key] is None:
                        del data[key]
                print(data)
                # 基于操作行为和对应参数来执行自动化操作
                # 实例化
                if values[1] == 'open_browser':
                    keys = Keys(**data)
                # 断言
                elif 'assert' in values[1]:
                    status = getattr(keys, values[1])(**data)
                    if status:
                        sheet.cell(row=values[0] + 2, column=8).value = 'Pass'
                    else:
                        sheet.cell(row=values[0] + 2, column=8).value = 'Failed'
                # 执行操作
                else:
                    getattr(keys, values[1])(**data)

    excel.save(file)
